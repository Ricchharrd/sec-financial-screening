from __future__ import annotations

from pathlib import Path


def autosize_worksheet(ws):
    from openpyxl.utils import get_column_letter

    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), 40)


def apply_header_style(ws):
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def append_sheet(ws, rows: list[dict], ratio_columns: set[str] | None = None, amount_columns: set[str] | None = None):
    ratio_columns = ratio_columns or set()
    amount_columns = amount_columns or set()
    if not rows:
        ws.append(["No data"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    apply_header_style(ws)
    autosize_worksheet(ws)
    for header in headers:
        col_idx = headers.index(header) + 1
        if header in ratio_columns:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = "0.0%"
        elif header in amount_columns:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row_idx, col_idx).number_format = '#,##0;[Red](#,##0);-'


def write_screening_workbook(
    summary_rows: list[dict],
    flag_rows: list[dict],
    note_rows: list[dict],
    output_path: Path,
    rating_rows: list[dict] | None = None,
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError("Excel export requires openpyxl. Install it with pip.") from exc

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws["A1"] = "SEC Preliminary Financial Screening"
    summary_ws["A1"].font = Font(bold=True, size=15)
    summary_ws["A2"] = "Notice"
    summary_ws["B2"] = "This is a preliminary screening tool, not a rating opinion. Human review required."

    append_sheet(
        summary_ws,
        summary_rows,
        ratio_columns={"Liabilities / Equity", "Operating Margin", "Net Margin", "ROA"},
        amount_columns={
            "Revenue",
            "Operating Income",
            "Net Income",
            "Total Assets",
            "Total Liabilities",
            "Total Equity",
            "Operating Cash Flow",
            "Working Capital",
        },
    )
    summary_ws.freeze_panes = "A4"

    flags_ws = wb.create_sheet("Red Flags")
    append_sheet(flags_ws, flag_rows)

    ratings_ws = wb.create_sheet("Internal Ratings")
    append_sheet(
        ratings_ws,
        rating_rows or [],
        amount_columns={"Value", "Points", "Weighted Points"},
    )

    notes_ws = wb.create_sheet("Notes")
    append_sheet(notes_ws, note_rows)

    wb.save(output_path)


def workbook_bytes(
    summary_rows: list[dict],
    flag_rows: list[dict],
    note_rows: list[dict],
    error_rows: list[dict],
    rating_rows: list[dict] | None = None,
) -> bytes:
    import tempfile

    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "sec_financial_screening.xlsx"
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Excel export requires openpyxl. Install it with pip.") from exc

        write_screening_workbook(summary_rows, flag_rows, note_rows, output_path, rating_rows=rating_rows)
        wb = load_workbook(output_path)
        if error_rows:
            errors_ws = wb.create_sheet("Errors")
            append_sheet(errors_ws, error_rows)
        wb.save(output_path)
        return output_path.read_bytes()
